"""
Karacabey Merinos — Fenotipik İndeks Hesaplama (Streamlit)
============================================================
Ham veri yükle → Veri kalite kontrol → Fenotipik indeks hesapla → Excel indir

Çalıştırma:
    streamlit run app.py
"""

import io
import re
import unicodedata
from datetime import datetime

import numpy as np
import pandas as pd
import streamlit as st
from openpyxl import load_workbook


# ═══════════════════════════════════════════════════════════════════════════
#  SABİT TANIMLAR
# ═══════════════════════════════════════════════════════════════════════════

REQUIRED_COLUMNS = {
    'koyun_irki':  ['koyun ırkı', 'koyun irki', 'ırk', 'irk'],
    'koyun_no':    ['koyun no', 'koyunno'],
    'koc_no':      ['koç no', 'kocno'],
    'koyun_yas':   ['koyun yaş', 'koyunyas'],
    'dogum_tipi':  ['doğum tipi', 'dogumtipi'],
    'cinsiyet':    ['cinsiyeti'],
    'ulusal_no':   ['ulusal no'],
    'dogum_ag':    ['doğum ağırlığı'],
    'canli_ag':    ['canlı ağırlık'],
    'kuzu_dogum':  ['kuzu doğum tarih'],
    'asim_tarih':  ['koç aşım tarihi'],
    'olcum_tarih': ['ölçüm ve tartım tarihi', 'tartım tarihi', 'sütten kesim tartım'],
    'durumu':      ['durumu:', 'durumu'],
    'durum_tarih': ['durum tarih'],
    'olum_yas':    ['kuzunun öldüğü yaş', 'öldüğü yaş', 'ölüm yaşı'],
    'gebelik':     ['gebelik süresi'],
    'grup_no':     ['tarih sıralı', 'grup no', 'kızgınlık grubu', 'aşım grubu', 'doğum grubu'],
}

# Durum sütununda ölüm anlamına gelen değerler (tam eşitlik, küçük harf + sade)
OLUM_KEYWORDS = {
    'ölüm', 'ölü', 'öldü', 'ölmüş', 'oldu',
    'olum', 'olu', 'olmus', 'olmuş',
}


def _strip_accents(s: str) -> str:
    """Türkçe karakterleri ASCII'ye sadeleştir (i̇→i, ü→u, ç→c, ş→s, ğ→g, ö→o, ı→i)."""
    s = ''.join(c for c in unicodedata.normalize('NFKD', s) if not unicodedata.combining(c))
    return s


# Doğum tipi normalize tablosu
DOGUM_TIPI_MAP = {
    # T - Tekiz
    't': 'T', 'tek': 'T', 'tekiz': 'T', 'tekli': 'T', 'single': 'T',
    # İ - İkiz
    'i': 'İ', 'ikiz': 'İ', 'ikili': 'İ', 'twin': 'İ',
    # Ü - Üçüz
    'u': 'Ü', 'ucuz': 'Ü', 'uc': 'Ü', 'uclu': 'Ü', 'triplet': 'Ü',
    # 4 - Dördüz
    '4': '4', 'dort': '4', 'dortuz': '4', 'dortlu': '4', 'quad': '4', 'quadruplet': '4',
}


def normalize_dogum_tipi(v):
    """Doğum tipini standart T/İ/Ü/4 formatına çevir (tek/ikiz/üçüz/dört vb. yakalanır)."""
    if pd.isna(v):
        return ''
    s = str(v).lower().strip()
    if not s:
        return ''
    s = _strip_accents(s)
    if s in DOGUM_TIPI_MAP:
        return DOGUM_TIPI_MAP[s]
    # Prefix match
    for prefix, code in (('tek', 'T'), ('iki', 'İ'), ('uc', 'Ü'), ('dor', '4')):
        if s.startswith(prefix):
            return code
    return str(v).strip()  # bilinmeyen → raw


def is_olum(v) -> bool:
    """Durum sütununda ölüm anlamı taşıyan değerleri yakala (esnek)."""
    if pd.isna(v):
        return False
    s = str(v).lower().strip().rstrip(':').strip()
    if not s:
        return False
    if s in OLUM_KEYWORDS:
        return True
    # 'öl' ile başlayan kısa kelimeler (ölüm, ölü, öldü, ölmüş vb.)
    if s.startswith('öl') and len(s) <= 8:
        return True
    return False

# AYARLAR sayfasından okunan varsayılan değerler (Is_Akisi_Sablonu_Final.xlsm)
DEFAULTS = {
    # Kriter ağırlıkları (toplam 100)
    'w_dkv':        17.0,   # DKV - Doğum Tipi
    'w_90kv':       25.0,   # 90KV - Kuzu Sayısı
    'w_top90ca':    20.0,   # Toplam 90CA
    'w_ortcaa':     12.0,   # OrtCAA
    'w_kizginlik':  14.0,   # Erken Kızgınlık (Aşım Grubu)
    'w_embriyo':    12.0,   # Embriyonik Kayıp
    # Toplam 90CA eşikleri
    't90_low':      27.41,
    't90_mid':      50.0,
    # OrtCAA eşikleri
    'caa_zero':     0.112,
    'caa_low':      0.311,
    'caa_mid':      0.35,
    # Erken Kızgınlık puanı için grup kaynağı: 'auto_dogum', 'grup_no', 'asim', 'dogum'
    'kizginlik_kaynak': 'auto_dogum',
    # Aşım/Doğum tarihinden grup hesaplama eşikleri (ISO hafta no)
    'w_grup1_max': 23,
    'w_grup2_eq':  24,
    'w_grup3_max': 26,
}


# ═══════════════════════════════════════════════════════════════════════════
#  YARDIMCI FONKSİYONLAR
# ═══════════════════════════════════════════════════════════════════════════

def find_columns(df: pd.DataFrame) -> dict:
    """Sütun isimlerini akıllı eşleştirme ile bulur. Çift boşluk/newline tolere edilir."""
    cm = {}
    used = set()
    for key, patterns in REQUIRED_COLUMNS.items():
        for col in df.columns:
            if col in used:
                continue
            # Boşlukları normalize et (\n, çift boşluk → tek boşluk)
            col_clean = re.sub(r'\s+', ' ', str(col).lower().strip())
            for pat in patterns:
                pat_clean = re.sub(r'\s+', ' ', pat.lower().strip())
                if pat_clean in col_clean and key not in cm:
                    if key == 'canli_ag' and ('1.' in col_clean or 'artış' in col_clean):
                        continue
                    if key == 'koyun_no' and ('koç' in col_clean or 'koc' in col_clean):
                        continue
                    cm[key] = col
                    used.add(col)
                    break
    return cm


def validate_data(df: pd.DataFrame, cm: dict) -> list[dict]:
    """Sütun kayma + format uyarıları."""
    warnings = []
    if 'koyun_no' in cm:
        sample = df[cm['koyun_no']].dropna().head(20)
        non_numeric = sum(1 for v in sample if not _is_numeric(v))
        if non_numeric > 5:
            warnings.append({
                'severity': 'high',
                'msg': f"'{cm['koyun_no']}' sütunu sayısal değil — sütun kayması olabilir!",
            })
    if 'dogum_tipi' in cm:
        vals = df[cm['dogum_tipi']].dropna().astype(str).str.strip().unique()
        valid = {'T', 'İ', 'Ü', '4'}
        invalid = [v for v in vals if v not in valid]
        if len(invalid) > 0:
            warnings.append({
                'severity': 'medium' if len(invalid) <= 3 else 'high',
                'msg': f"'{cm['dogum_tipi']}' sütununda beklenmeyen değerler: {invalid[:5]}",
            })
    if 'cinsiyet' in cm:
        vals = df[cm['cinsiyet']].dropna().astype(str).str.strip().unique()
        invalid = [v for v in vals if not v or v[0] not in ('E', 'D', 'Ö')]
        if len(invalid) > 3:
            warnings.append({
                'severity': 'medium',
                'msg': f"'{cm['cinsiyet']}' sütununda beklenmeyen değerler: {invalid[:5]}",
            })
    if 'dogum_ag' in cm:
        with_comma = df[cm['dogum_ag']].astype(str).str.contains(',', na=False).sum()
        if with_comma > 0:
            warnings.append({
                'severity': 'info',
                'msg': f"'{cm['dogum_ag']}' sütununda {with_comma} virgüllü değer var — otomatik dönüştürülecek.",
            })
    return warnings


def _is_numeric(v) -> bool:
    try:
        float(v)
        return True
    except (TypeError, ValueError):
        return False


def quality_report(df: pd.DataFrame, cm: dict, olum_checker=None) -> dict:
    """'2-Veri Kalite Kontrol' sayfasının Python karşılığı."""
    if olum_checker is None:
        olum_checker = is_olum
    r = {}
    if 'koyun_no' in cm:
        kn = df[cm['koyun_no']].dropna()
        r['toplam_kayit'] = len(kn)
        r['benzersiz_koyun'] = kn.nunique()
    if 'ulusal_no' in cm and 'cinsiyet' in cm:
        olu = df[cm['cinsiyet']].astype(str).str.upper().str.strip() == 'ÖLÜ'
        r['benzersiz_kuzu'] = df.loc[~olu, cm['ulusal_no']].dropna().nunique()
        r['olu_dogum'] = int(olu.sum())
    # Doğum tipi dağılımı
    if 'dogum_tipi' in cm:
        dt = df[cm['dogum_tipi']].astype(str).str.strip()
        r['dt_T'] = int((dt == 'T').sum())
        r['dt_İ'] = int((dt == 'İ').sum())
        r['dt_Ü'] = int((dt == 'Ü').sum())
        r['dt_4'] = int((dt == '4').sum())
    # Eksik veri
    for k, label in [('koc_no', 'eksik_koc'), ('koyun_no', 'eksik_koyun'),
                     ('dogum_ag', 'eksik_dogag'), ('canli_ag', 'eksik_canag'),
                     ('olcum_tarih', 'eksik_olctar')]:
        if k in cm:
            r[label] = int(df[cm[k]].isna().sum())
    # Ölüm istatistikleri
    if 'durumu' in cm:
        ol = df[cm['durumu']].apply(olum_checker)
        r['toplam_olum'] = int(ol.sum())
        # Ölüm yaşı: olum_yas varsa ve dolu ise onu kullan; aksi halde durum_tarih - kuzu_dogum
        oy = None
        if 'olum_yas' in cm:
            candidate = pd.to_numeric(df.loc[ol, cm['olum_yas']], errors='coerce')
            if candidate.notna().any():
                oy = candidate
        if oy is None and 'durum_tarih' in cm and 'kuzu_dogum' in cm:
            d_dur = pd.to_datetime(df.loc[ol, cm['durum_tarih']], errors='coerce')
            d_dog = pd.to_datetime(df.loc[ol, cm['kuzu_dogum']], errors='coerce')
            oy = (d_dur - d_dog).dt.days
        if oy is not None and len(oy) > 0:
            r['ort_olum_yas'] = float(oy.mean()) if oy.notna().any() else None
            r['olum_90_oncesi'] = int((oy < 90).sum())
            r['olum_90_sonrasi'] = int((oy >= 90).sum())
    return r


def clean_data(df: pd.DataFrame, cm: dict, P: dict | None = None, olum_checker=None) -> pd.DataFrame:
    """Virgül düzeltme, 90g ağırlık hesabı, aşım grubu, yaşama gücü."""
    if olum_checker is None:
        olum_checker = is_olum
    if P is None:
        P = DEFAULTS
    df = df.copy()
    for ck in ('dogum_ag', 'canli_ag'):
        if ck in cm:
            df[cm[ck]] = (df[cm[ck]].astype(str).str.replace(',', '.', regex=False)
                                                  .replace({'nan': np.nan, 'None': np.nan}))
            df[cm[ck]] = pd.to_numeric(df[cm[ck]], errors='coerce')

    if 'gebelik' in cm:
        gc = cm['gebelik']
        df[gc] = pd.to_numeric(df[gc], errors='coerce')
        m = df[gc] > 200
        if m.any():
            df.loc[m, gc] = df.loc[m, gc].apply(
                lambda x: int(str(int(x))[:3]) if pd.notna(x) else x
            )

    if 'cinsiyet' in cm:
        df['_olu'] = df[cm['cinsiyet']].astype(str).str.upper().str.strip().str.startswith('ÖLÜ')
    else:
        df['_olu'] = False

    if 'koc_no' in cm:
        df[cm['koc_no']] = df[cm['koc_no']].fillna(0)

    if 'kuzu_dogum' in cm and 'olcum_tarih' in cm:
        df['_ty'] = (pd.to_datetime(df[cm['olcum_tarih']], errors='coerce')
                     - pd.to_datetime(df[cm['kuzu_dogum']], errors='coerce')).dt.days

    if 'dogum_ag' in cm and 'canli_ag' in cm:
        d = df[cm['dogum_ag']]
        c2 = df[cm['canli_ag']]
        y = df.get('_ty', pd.Series([np.nan] * len(df)))
        df['_90g'] = np.where((y > 0) & c2.notna() & d.notna(),
                              d + ((c2 - d) / y) * 90, np.nan)

    if 'durumu' in cm:
        ol = df[cm['durumu']].apply(olum_checker)
        # Ölüm yaşı: olum_yas dolu ise onu kullan; aksi halde durum_tarih - kuzu_dogum
        oy = None
        if 'olum_yas' in cm:
            candidate = pd.to_numeric(df[cm['olum_yas']], errors='coerce')
            if candidate.notna().any():
                oy = candidate
        if oy is None and 'durum_tarih' in cm and 'kuzu_dogum' in cm:
            d_dur = pd.to_datetime(df[cm['durum_tarih']], errors='coerce')
            d_dog = pd.to_datetime(df[cm['kuzu_dogum']], errors='coerce')
            oy = (d_dur - d_dog).dt.days
        if oy is None:
            oy = pd.Series([np.nan] * len(df))
        df['_olum_yas'] = oy
        df['_yas_gucu'] = np.where(df['_olu'], 0,
                          np.where(ol & (oy < 90), 0,
                          np.where(ol & (oy >= 90), 1,
                          np.where(~ol, 1, np.nan))))
    else:
        df['_olum_yas'] = np.nan
        df['_yas_gucu'] = np.where(df['_olu'], 0, 1)

    g1 = int(P.get('w_grup1_max', 23))
    g2 = int(P.get('w_grup2_eq', 24))
    g3 = int(P.get('w_grup3_max', 26))
    def grup(h):
        if pd.isna(h):
            return np.nan
        if h <= g1:
            return 1
        if h == g2:
            return 2
        if h <= g3:
            return 3
        return 4

    if 'asim_tarih' in cm:
        hafta_a = pd.to_datetime(df[cm['asim_tarih']], errors='coerce').dt.isocalendar().week.astype('Int64')
        df['_asim_grup'] = hafta_a.apply(grup)
    if 'kuzu_dogum' in cm:
        hafta_d = pd.to_datetime(df[cm['kuzu_dogum']], errors='coerce').dt.isocalendar().week.astype('Int64')
        df['_dogum_grup'] = hafta_d.apply(grup)
    if 'grup_no' in cm:
        df['_grup_no'] = pd.to_numeric(df[cm['grup_no']], errors='coerce')
    return df


def add_auto_grup(piv: pd.DataFrame, n_groups: int = 3) -> pd.DataFrame:
    """Koyunları İlk Doğum Tarihine göre n_groups eşit parçaya böl (tertile).
    En erken doğuranlar → Grup 1, geç doğuranlar → Grup n_groups.
    Tarih yoksa veya geçersizse → Grup (n_groups + 1) = "doğum yok" / 4.
    """
    if 'İlk Doğum Tarihi' not in piv.columns:
        piv['Otomatik Grup'] = pd.Series([pd.NA] * len(piv), dtype='Int64')
        return piv
    tarihler = pd.to_datetime(piv['İlk Doğum Tarihi'], errors='coerce')
    valid = tarihler.notna()
    out = pd.Series([n_groups + 1] * len(piv), index=piv.index, dtype='Int64')
    if valid.sum() >= n_groups:
        # rank ile aynı tarihleri sıralamada stabil tut, sonra eşit parçalara böl
        ranks = tarihler[valid].rank(method='first')
        cuts = pd.qcut(ranks, q=n_groups, labels=list(range(1, n_groups + 1)))
        out.loc[valid] = cuts.astype(int).astype('Int64')
    elif valid.sum() > 0:
        # n_groups'tan az koyun varsa hepsini Grup 1 yap
        out.loc[valid] = 1
    piv['Otomatik Grup'] = out
    return piv


def build_pivot(df: pd.DataFrame, cm: dict, olum_checker=None) -> pd.DataFrame:
    """Koyun bazında pivot — her koyun için max 3 kuzu (K1/K2/K3)."""
    if olum_checker is None:
        olum_checker = is_olum
    df_a = df[~df['_olu']].copy()
    rows = []
    for kn, g in df_a.groupby(cm['koyun_no']):
        try:
            kn_int = int(float(kn))
        except (TypeError, ValueError):
            continue
        r = {'Koyun No': kn_int}
        r['Koyun Yaş'] = g[cm['koyun_yas']].iloc[0] if 'koyun_yas' in cm else None
        r['Doğum Tipi'] = str(g[cm['dogum_tipi']].iloc[0]).strip() if 'dogum_tipi' in cm else ''
        for i in range(3):
            ki = i + 1
            if i < len(g):
                ku = g.iloc[i]
                r[f'K{ki} Ulusal'] = ku.get(cm.get('ulusal_no', ''), '')
                r[f'K{ki} DOGAG'] = ku.get(cm.get('dogum_ag', ''), np.nan)
                r[f'K{ki} CA90g'] = ku.get('_90g', np.nan)
                dur_val = ku.get(cm.get('durumu', ''), '')
                oy_num = ku.get('_olum_yas', np.nan)
                try:
                    oy_f = float(oy_num) if pd.notna(oy_num) else None
                except (TypeError, ValueError):
                    oy_f = None
                olum = 'Kuzu ölümü' if olum_checker(dur_val) and oy_f is not None and oy_f < 90 else ''
                r[f'K{ki} Ölüm'] = olum
                r[f'K{ki} YasamaGucu'] = ku.get('_yas_gucu', np.nan)
            else:
                r[f'K{ki} Ulusal'] = ''
                r[f'K{ki} DOGAG'] = np.nan
                r[f'K{ki} CA90g'] = np.nan
                r[f'K{ki} Ölüm'] = ''
                r[f'K{ki} YasamaGucu'] = np.nan
        r['Aşım Grubu']  = g['_asim_grup'].iloc[0]  if '_asim_grup'  in g.columns else np.nan
        r['Doğum Grubu'] = g['_dogum_grup'].iloc[0] if '_dogum_grup' in g.columns else np.nan
        r['Grup No']     = g['_grup_no'].iloc[0]    if '_grup_no'    in g.columns else np.nan
        # İlk doğum tarihini (en eski) bul — otomatik grup için
        if 'kuzu_dogum' in cm:
            try:
                r['İlk Doğum Tarihi'] = pd.to_datetime(g[cm['kuzu_dogum']], errors='coerce').min()
            except Exception:
                r['İlk Doğum Tarihi'] = pd.NaT
        else:
            r['İlk Doğum Tarihi'] = pd.NaT
        rows.append(r)
    piv = pd.DataFrame(rows)
    piv = add_auto_grup(piv, n_groups=3)
    return piv


def calc_index(piv: pd.DataFrame, P: dict) -> pd.DataFrame:
    """Fenotipik indeks hesaplama — Excel formüllerinin Python karşılığı."""
    df = piv.copy()

    # P — Kuzu Ölüm Sayısı
    df['Ölüm Sayısı'] = sum(
        (df[f'K{i} Ölüm'].astype(str) == 'Kuzu ölümü').astype(int) for i in (1, 2, 3)
    )

    # Q/R/S — günlük canlı ağırlık artışı (CAA)
    for i in (1, 2, 3):
        ca = pd.to_numeric(df[f'K{i} CA90g'], errors='coerce')
        dg = pd.to_numeric(df[f'K{i} DOGAG'], errors='coerce')
        df[f'K{i} CAA'] = np.where(ca > 0, (ca - dg) / 90, 0.0)

    # U — Doğan Kuzu Sayısı (CA90g > 0 olanlar)
    df['Kuzu Sayısı'] = sum(
        (pd.to_numeric(df[f'K{i} CA90g'], errors='coerce') > 0).astype(int) for i in (1, 2, 3)
    )

    # V — Toplam 90CA
    df['Toplam 90CA'] = sum(
        pd.to_numeric(df[f'K{i} CA90g'], errors='coerce').fillna(0) for i in (1, 2, 3)
    )

    # T — OrtCAA (Kuzu sayısına göre çarpan: 1→×1, 2→×1.1, 3→×1.2)
    def ortcaa(row):
        u = row['Kuzu Sayısı']
        if u == 0:
            return 0.0
        vals = [row[f'K{i} CAA'] for i in (1, 2, 3) if row[f'K{i} CAA'] > 0]
        if not vals:
            return 0.0
        avg = sum(vals) / len(vals)
        if u == 1:
            return avg
        if u == 2:
            return avg * 1.1
        if u == 3:
            return avg * 1.2
        return 0.0
    df['OrtCAA'] = df.apply(ortcaa, axis=1)

    # W — Toplam 90CA katsayısı
    def w_coef(v):
        if v == 0:
            return 0.0
        if 0 < v <= P['t90_low']:
            return 0.5
        if P['t90_low'] < v <= P['t90_mid']:
            return 0.85
        if v > P['t90_mid']:
            return 1.0
        return 0.0
    df['Toplam 90CA Kats.'] = df['Toplam 90CA'].apply(w_coef)

    # X — OrtCAA katsayısı
    def x_coef(v):
        if v == 0 or v <= P['caa_zero']:
            return 0.0
        if P['caa_zero'] < v <= P['caa_low']:
            return 0.5
        if P['caa_low'] < v <= P['caa_mid']:
            return 0.85
        if v > P['caa_mid']:
            return 1.0
        return 0.0
    df['OrtCAA Kats.'] = df['OrtCAA'].apply(x_coef)

    # Y — DKV puanı (Doğum Tipi)
    def dkv(t):
        t = str(t).strip()
        if t == 'T':
            return P['w_dkv'] * 0.3
        if t == 'İ':
            return P['w_dkv'] * 0.7
        if t == 'Ü':
            return P['w_dkv'] * 1.0
        return 0.0
    df['DKV Puan'] = df['Doğum Tipi'].apply(dkv)

    # Z — 90KV puanı (Kuzu Sayısı)
    def kv90(u):
        if u == 1:
            return P['w_90kv'] * 0.7
        if u == 2:
            return P['w_90kv'] * 0.9
        if u == 3:
            return P['w_90kv'] * 1.0
        return 0.0
    df['90KV Puan'] = df['Kuzu Sayısı'].apply(kv90)

    # AA — Toplam 90CA puanı
    df['Toplam 90CA Puan'] = P['w_top90ca'] * df['Toplam 90CA Kats.']

    # AB — OrtCAA puanı
    df['OrtCAA Puan'] = P['w_ortcaa'] * df['OrtCAA Kats.']

    # AC — Kriter Puan Toplamı
    df['Kriter Toplam'] = df[['DKV Puan', '90KV Puan',
                              'Toplam 90CA Puan', 'OrtCAA Puan']].sum(axis=1)

    # AE — Erken Kızgınlık puanı (Aşım veya Doğum grubuna göre — sidebar seçimi)
    def kizg(g):
        if pd.isna(g):
            return 0.0
        g = int(g)
        if g == 1:
            return P['w_kizginlik'] * 1.0
        if g == 2:
            return P['w_kizginlik'] * 0.85
        if g == 3:
            return P['w_kizginlik'] * 0.3
        return 0.0
    kaynak = P.get('kizginlik_kaynak', 'asim')
    grup_kol_map = {
        'asim': 'Aşım Grubu',
        'dogum': 'Doğum Grubu',
        'grup_no': 'Grup No',
        'auto_dogum': 'Otomatik Grup',
    }
    grup_kol = grup_kol_map.get(kaynak, 'Aşım Grubu')
    # Seçilen kolon yok veya tamamen NaN ise fallback
    if grup_kol not in df.columns or df[grup_kol].notna().sum() == 0:
        for alt in ('Otomatik Grup', 'Grup No', 'Aşım Grubu', 'Doğum Grubu'):
            if alt in df.columns and df[alt].notna().sum() > 0:
                grup_kol = alt
                break
    df['Erken Kızgınlık Puan'] = df[grup_kol].apply(kizg) if grup_kol in df.columns else 0.0
    df['Erken Kızgınlık Kaynak'] = grup_kol

    # AF — Embriyonik Kayıp puanı (varsayılan: Doğum var=1.0)
    df['Embriyonik Kayıp Puan'] = P['w_embriyo'] * 1.0

    # AG — Nihai Fenotipik İndeks
    df['Fenotipik İndeks'] = (df['Kriter Toplam']
                              + df['Erken Kızgınlık Puan']
                              + df['Embriyonik Kayıp Puan']).round(2)

    # Görüntü sırası
    cols = [
        'Koyun No', 'Koyun Yaş', 'Doğum Tipi',
        'K1 Ulusal', 'K1 DOGAG', 'K1 CA90g', 'K1 CAA', 'K1 Ölüm', 'K1 YasamaGucu',
        'K2 Ulusal', 'K2 DOGAG', 'K2 CA90g', 'K2 CAA', 'K2 Ölüm', 'K2 YasamaGucu',
        'K3 Ulusal', 'K3 DOGAG', 'K3 CA90g', 'K3 CAA', 'K3 Ölüm', 'K3 YasamaGucu',
        'İlk Doğum Tarihi',
        'Aşım Grubu', 'Doğum Grubu', 'Grup No', 'Otomatik Grup',
        'Kuzu Sayısı', 'Ölüm Sayısı',
        'Toplam 90CA', 'Toplam 90CA Kats.', 'OrtCAA', 'OrtCAA Kats.',
        'DKV Puan', '90KV Puan', 'Toplam 90CA Puan', 'OrtCAA Puan',
        'Kriter Toplam', 'Erken Kızgınlık Puan', 'Erken Kızgınlık Kaynak',
        'Embriyonik Kayıp Puan',
        'Fenotipik İndeks',
    ]
    return df[cols]


def to_excel_bytes(result: pd.DataFrame, params: dict, kalite: dict) -> bytes:
    """Tüm tabloları tek bir Excel dosyasına yazar."""
    bio = io.BytesIO()
    with pd.ExcelWriter(bio, engine='xlsxwriter') as xw:
        result.to_excel(xw, sheet_name='Fenotipik Indeks', index=False)
        # Ayarlar snapshot
        params_df = pd.DataFrame([
            ('DKV (Doğum Tipi) Ağırlık',     params['w_dkv']),
            ('90KV (Kuzu Sayısı) Ağırlık',   params['w_90kv']),
            ('Toplam 90CA Ağırlık',          params['w_top90ca']),
            ('OrtCAA Ağırlık',               params['w_ortcaa']),
            ('Erken Kızgınlık Ağırlık',      params['w_kizginlik']),
            ('Embriyonik Kayıp Ağırlık',     params['w_embriyo']),
            ('Toplam 90CA - Alt Eşik',       params['t90_low']),
            ('Toplam 90CA - Üst Eşik',       params['t90_mid']),
            ('OrtCAA - Sıfır Eşik',          params['caa_zero']),
            ('OrtCAA - Alt Eşik',            params['caa_low']),
            ('OrtCAA - Üst Eşik',            params['caa_mid']),
        ], columns=['Parametre', 'Değer'])
        params_df.to_excel(xw, sheet_name='Ayarlar', index=False)
        # Kalite raporu
        if kalite:
            kalite_df = pd.DataFrame(list(kalite.items()), columns=['Kontrol', 'Değer'])
            kalite_df.to_excel(xw, sheet_name='Veri Kalite', index=False)
        # Sütunları otomatik genişlik
        for sh, df_ in [('Fenotipik Indeks', result), ('Ayarlar', params_df)]:
            ws = xw.sheets[sh]
            for i, col in enumerate(df_.columns):
                if len(df_):
                    try:
                        data_max = max((len(str(v)) for v in df_[col].tolist()), default=12)
                    except Exception:
                        data_max = 12
                else:
                    data_max = 12
                header_len = len(str(col)) + 2
                width = max(12, min(28, max(data_max, header_len)))
                ws.set_column(i, i, width)
    return bio.getvalue()


def read_ayarlar_from_template(file_bytes: bytes) -> dict | None:
    """Yüklenen Excel dosyasının AYARLAR sayfasından parametreleri oku."""
    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True, keep_vba=False)
        if 'AYARLAR' not in wb.sheetnames:
            return None
        ws = wb['AYARLAR']
        P = dict(DEFAULTS)
        def num(coord):
            v = ws[coord].value
            try:
                return float(v)
            except (TypeError, ValueError):
                return None
        for key, coord in [
            ('w_dkv', 'B5'), ('w_90kv', 'B6'), ('w_top90ca', 'B7'),
            ('w_ortcaa', 'B8'), ('w_kizginlik', 'B9'), ('w_embriyo', 'B10'),
            ('t90_low', 'C16'), ('t90_mid', 'C17'),
            ('caa_zero', 'C22'), ('caa_low', 'C23'), ('caa_mid', 'C24'),
        ]:
            v = num(coord)
            if v is not None:
                P[key] = v
        return P
    except Exception:
        return None


# ═══════════════════════════════════════════════════════════════════════════
#  STREAMLIT UI
# ═══════════════════════════════════════════════════════════════════════════

st.set_page_config(page_title='Fenotipik İndeks Hesaplama', layout='wide')

st.title('Fenotipik İndeks Hesaplama')
st.caption('Karacabey Merinos — Ham veri → Veri kalite kontrol → Fenotipik indeks')

# Session state
if 'params' not in st.session_state:
    st.session_state.params = dict(DEFAULTS)
if 'step' not in st.session_state:
    st.session_state.step = 1

# ──────────────────────────────────────────────────────────────────────────
# SIDEBAR — AYARLAR
# ──────────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.header('AYARLAR')

    with st.expander('Şablon dosyadan AYARLAR yükle', expanded=False):
        ayar_file = st.file_uploader('Excel (.xlsx/.xlsm)', type=['xlsx', 'xlsm'], key='ayar_file')
        if ayar_file is not None:
            P_loaded = read_ayarlar_from_template(ayar_file.getvalue())
            if P_loaded:
                st.session_state.params = P_loaded
                st.success('AYARLAR sayfasından okundu.')
            else:
                st.warning('AYARLAR sayfası bulunamadı veya okunamadı.')

    P = st.session_state.params

    st.subheader('1. Kriter Ağırlıkları')
    c1, c2 = st.columns(2)
    P['w_dkv']       = c1.number_input('DKV (Doğum Tipi)',  value=float(P['w_dkv']),       step=1.0, min_value=0.0)
    P['w_90kv']      = c2.number_input('90KV (Kuzu Sayısı)', value=float(P['w_90kv']),     step=1.0, min_value=0.0)
    P['w_top90ca']   = c1.number_input('Toplam 90CA',        value=float(P['w_top90ca']),   step=1.0, min_value=0.0)
    P['w_ortcaa']    = c2.number_input('OrtCAA',             value=float(P['w_ortcaa']),    step=1.0, min_value=0.0)
    P['w_kizginlik'] = c1.number_input('Erken Kızgınlık',    value=float(P['w_kizginlik']), step=1.0, min_value=0.0)
    P['w_embriyo']   = c2.number_input('Embriyonik Kayıp',   value=float(P['w_embriyo']),   step=1.0, min_value=0.0)

    toplam = (P['w_dkv'] + P['w_90kv'] + P['w_top90ca']
              + P['w_ortcaa'] + P['w_kizginlik'] + P['w_embriyo'])
    if abs(toplam - 100) < 0.01:
        st.success(f'Toplam ağırlık: {toplam:.1f} ✓')
    else:
        st.warning(f'Toplam ağırlık: {toplam:.1f} (100 olmalı)')

    st.subheader('2. Toplam 90CA Eşikleri')
    P['t90_low'] = st.number_input('Alt eşik (0.5 puan üstü)', value=float(P['t90_low']), step=0.1)
    P['t90_mid'] = st.number_input('Üst eşik (1.0 puan üstü)', value=float(P['t90_mid']), step=0.1)

    st.subheader('3. OrtCAA Eşikleri')
    P['caa_zero'] = st.number_input('Sıfır eşik',           value=float(P['caa_zero']), step=0.001, format='%.3f')
    P['caa_low']  = st.number_input('Alt eşik (0.5 üstü)',  value=float(P['caa_low']),  step=0.001, format='%.3f')
    P['caa_mid']  = st.number_input('Üst eşik (1.0 üstü)',  value=float(P['caa_mid']),  step=0.001, format='%.3f')

    st.subheader('4. Erken Kızgınlık Kaynağı')
    kaynak_opts = ['auto_dogum', 'grup_no', 'asim', 'dogum']
    kaynak_labels = {
        'auto_dogum': 'Otomatik 3\'lü Bölme (doğum sırasından — önerilen)',
        'grup_no':    'Hazır Grup No sütunu (1/2/3)',
        'asim':       'Koç Aşım Tarihi (ISO haftadan)',
        'dogum':      'Kuzu Doğum Tarihi (ISO haftadan)',
    }
    cur = P.get('kizginlik_kaynak', 'auto_dogum')
    if cur not in kaynak_opts:
        cur = 'auto_dogum'
    P['kizginlik_kaynak'] = st.radio(
        'Erken Kızgınlık puanı hangi gruptan hesaplansın?',
        options=kaynak_opts,
        index=kaynak_opts.index(cur),
        format_func=lambda x: kaynak_labels[x],
        help='"Hazır Grup No" seçersen, eşleştirmede "Grup No" alanını da seçmen gerekir (örn. "tarih sıralı" sütunu).',
    )

    st.subheader('5. Grup Hafta Eşikleri (Aşım/Doğum tarihinden)')
    st.caption('"Hazır Grup No" seçtiysen bu kısım kullanılmaz.')
    P['w_grup1_max'] = st.number_input(
        'Grup 1 üst sınır (hafta ≤)',
        value=int(P.get('w_grup1_max', 23)), min_value=1, max_value=53, step=1,
    )
    P['w_grup2_eq']  = st.number_input(
        'Grup 2 hafta (=)',
        value=int(P.get('w_grup2_eq', 24)), min_value=1, max_value=53, step=1,
    )
    P['w_grup3_max'] = st.number_input(
        'Grup 3 üst sınır (hafta ≤)',
        value=int(P.get('w_grup3_max', 26)), min_value=1, max_value=53, step=1,
    )

    if st.button('Varsayılana sıfırla'):
        st.session_state.params = dict(DEFAULTS)
        st.rerun()

# ──────────────────────────────────────────────────────────────────────────
# ADIM 1 — DOSYA YÜKLE
# ──────────────────────────────────────────────────────────────────────────
st.header('1. Ham Veri Yükle')
data_file = st.file_uploader(
    'Dosya seç (.xlsx / .xlsm / .csv)',
    type=['xlsx', 'xlsm', 'csv'],
    key='data_file',
    help="Excel dosyası veya CSV. Şablon dosyası kullanıyorsan '1-Ham Veri Yapıştır' sayfasını seç.",
)

if data_file is None:
    st.info('Devam etmek için bir dosya yükle.')
    st.stop()

is_csv = data_file.name.lower().endswith('.csv')

df_raw = None

if is_csv:
    # --- CSV ---
    c_enc, c_sep, c_hdr = st.columns(3)
    encoding = c_enc.selectbox(
        'Kodlama (encoding)',
        ['utf-8', 'utf-8-sig', 'cp1254', 'latin1'],
        index=0,
        help='Türkçe karakterlerde sorun varsa cp1254 dene.',
    )
    sep_choice = c_sep.selectbox(
        'Ayraç (separator)',
        ['Otomatik', 'Virgül (,)', 'Noktalı virgül (;)', 'Tab', 'Boru (|)'],
        index=0,
    )
    sep_map = {'Otomatik': None, 'Virgül (,)': ',', 'Noktalı virgül (;)': ';', 'Tab': '\t', 'Boru (|)': '|'}
    sep = sep_map[sep_choice]
    header_row = c_hdr.number_input(
        'Başlık satırı (0 = başlık yok)',
        min_value=0, max_value=20, value=1, step=1,
        help='Başlıklar hangi satırda? 0 yazarsan başlıklar yok sayılır, sütunlar 0,1,2,... olur ve aşağıdan manuel eşleştirirsin.',
    )

    try:
        data_file.seek(0)
        df_raw = pd.read_csv(
            data_file,
            sep=sep,
            engine='python' if sep is None else 'c',
            header=(header_row - 1) if header_row > 0 else None,
            encoding=encoding,
        )
        df_raw.columns = [re.sub(r'\s+', ' ', str(c).replace('\n', ' ').strip()) for c in df_raw.columns]
        df_raw = df_raw.dropna(how='all').reset_index(drop=True)
    except Exception as e:
        st.error(f'CSV okunamadı: {e}')
        st.stop()

else:
    # --- Excel ---
    try:
        xl = pd.ExcelFile(data_file, engine='openpyxl')
    except Exception as e:
        st.error(f'Dosya okunamadı: {e}')
        st.stop()

    default_idx = 0
    for i, name in enumerate(xl.sheet_names):
        if 'ham veri' in name.lower():
            default_idx = i
            break

    c_sheet, c_header = st.columns([2, 1])
    sheet = c_sheet.selectbox('Sayfa', xl.sheet_names, index=default_idx)
    # Şablon sayfası seçildiyse varsayılan 3, normal sayfa için 1
    is_template = 'ham veri yapıştır' in sheet.lower()
    header_row = c_header.number_input(
        'Başlık satırı (1-tabanlı)',
        min_value=1, max_value=10,
        value=3 if is_template else 1,
        step=1,
        help='Normal Excel: 1. "1-Ham Veri Yapıştır" şablonu: 3.',
    )

    try:
        df_raw = pd.read_excel(data_file, sheet_name=sheet, header=header_row - 1, engine='openpyxl')
        df_raw.columns = [re.sub(r'\s+', ' ', str(c).replace('\n', ' ').strip()) for c in df_raw.columns]
        df_raw = df_raw.dropna(how='all').reset_index(drop=True)
    except Exception as e:
        st.error(f'Sayfa okunamadı: {e}')
        st.stop()

st.success(f'✓ {len(df_raw)} satır, {len(df_raw.columns)} sütun yüklendi')

with st.expander('Ham veri önizleme (ilk 20 satır)', expanded=False):
    st.dataframe(df_raw.head(20), use_container_width=True)

# ──────────────────────────────────────────────────────────────────────────
# ADIM 2 — VERİ KALİTE KONTROL
# ──────────────────────────────────────────────────────────────────────────
st.header('2. Veri Kalite Kontrol')

# Otomatik tanıma — varsayılan eşleşmeler
cm_auto = find_columns(df_raw)

# Alan etiketleri ve zorunlu olanlar
FIELD_LABELS = {
    'koyun_irki':  'Koyun Irkı (filtre için)',
    'koyun_no':    'Koyun No ★',
    'koc_no':      'Koç No',
    'koyun_yas':   'Koyun Yaş',
    'dogum_tipi':  'Doğum Tipi ★',
    'cinsiyet':    'Cinsiyet (E/D/ÖLÜ)',
    'ulusal_no':   'Ulusal No (kuzu)',
    'dogum_ag':    'Doğum Ağırlığı',
    'canli_ag':    'Canlı Ağırlık (son ölçüm)',
    'kuzu_dogum':  'Kuzu Doğum Tarihi',
    'asim_tarih':  'Koç Aşım Tarihi',
    'olcum_tarih': 'Ölçüm/Tartım Tarihi',
    'durumu':      'Durumu (ölüm/yaşıyor)',
    'durum_tarih': 'Durum Tarihi (ölüm tarihi)',
    'olum_yas':    'Ölüm Yaşı (gün) — boşsa durum tarihinden hesaplanır',
    'gebelik':     'Gebelik Süresi',
    'grup_no':     'Grup No (1/2/3) — "tarih sıralı" gibi hazır grup sütunu',
}

# Manuel sütun eşleştirme
st.subheader('Sütun Eşleştirme')
st.caption('Her alan için doğru sütunu seç. Otomatik tanınanlar önceden seçili — yanlışsa veya boşsa düzelt. ★ = zorunlu')

NONE_OPT = '(Yok)'
col_options = [NONE_OPT] + list(df_raw.columns)

# Session state ile seçimleri sakla (rerun'da kaybolmasın)
if 'col_mapping' not in st.session_state or st.session_state.get('_last_cols') != tuple(df_raw.columns):
    st.session_state.col_mapping = {k: cm_auto.get(k, NONE_OPT) for k in REQUIRED_COLUMNS}
    st.session_state._last_cols = tuple(df_raw.columns)

with st.expander('Eşleştirmeyi düzenle', expanded=True):
    cA, cB = st.columns(2)
    fields = list(REQUIRED_COLUMNS.keys())
    half = (len(fields) + 1) // 2
    for i, field in enumerate(fields):
        target = cA if i < half else cB
        with target:
            current = st.session_state.col_mapping.get(field, NONE_OPT)
            if current not in col_options:
                current = NONE_OPT
            sel = st.selectbox(
                FIELD_LABELS.get(field, field),
                col_options,
                index=col_options.index(current),
                key=f'col_{field}',
            )
            st.session_state.col_mapping[field] = sel

    if st.button('Otomatik tanıyı tekrar uygula'):
        for k in REQUIRED_COLUMNS:
            st.session_state.col_mapping[k] = cm_auto.get(k, NONE_OPT)
            st.session_state[f'col_{k}'] = st.session_state.col_mapping[k]
        st.rerun()

# Son cm — kullanıcı seçimlerinden
cm = {k: v for k, v in st.session_state.col_mapping.items() if v and v != NONE_OPT}

# Doğum Tipi normalize — "tek/ikiz/üçüz/dört" gibi açık ifadeleri T/İ/Ü/4'e çevir
if 'dogum_tipi' in cm:
    dt_col = cm['dogum_tipi']
    df_raw[dt_col] = df_raw[dt_col].apply(normalize_dogum_tipi)

# Eşleşme özeti
eslesen = len(cm)
eksik_zorunlu = [k for k in ('koyun_no', 'dogum_tipi') if k not in cm]
c1, c2, c3 = st.columns(3)
c1.metric('Eşleşen Alan', f'{eslesen}/{len(REQUIRED_COLUMNS)}')
c2.metric('Toplam Sütun', len(df_raw.columns))
c3.metric('Zorunlu Eksik', len(eksik_zorunlu))

if eksik_zorunlu:
    st.error(f"Zorunlu alanlar eksik: {[FIELD_LABELS[k] for k in eksik_zorunlu]}. Yukarıdan eşleştir.")

# Irk Filtresi — eşleşmişse, kullanıcı bir veya birden fazla ırk seçebilir
if 'koyun_irki' in cm:
    st.subheader('Irk Filtresi')
    irk_col = cm['koyun_irki']
    unique_irklar = sorted(
        {str(v).strip() for v in df_raw[irk_col].dropna().tolist() if str(v).strip()}
    )
    irk_sess_key = f'irk_filter_{irk_col}'
    if irk_sess_key not in st.session_state:
        st.session_state[irk_sess_key] = unique_irklar  # varsayılan: hepsi seçili

    selected_irklar = st.multiselect(
        f'"{irk_col}" sütununda hangi ırklar hesaplansın?',
        options=unique_irklar,
        key=irk_sess_key,
        help='Birden fazla ırk varsa sadece istediklerini seç. Boş bırakırsan hepsi dahil sayılır.',
    )
    if selected_irklar and len(selected_irklar) < len(unique_irklar):
        before = len(df_raw)
        df_raw = df_raw[df_raw[irk_col].astype(str).str.strip().isin(selected_irklar)].reset_index(drop=True)
        st.info(f"Irk filtresi: **{', '.join(selected_irklar)}** — {before} → {len(df_raw)} satır.")
    elif not selected_irklar:
        st.warning('Hiç ırk seçilmedi — tüm satırlar dahil edildi.')
    else:
        st.caption(f'Tüm ırklar dahil ({len(unique_irklar)} ırk, {len(df_raw)} satır).')

# Ölüm Değerleri — dinamik seçim (Durumu sütunundaki hangi değerler "kuzu ölümü"?)
olum_checker = is_olum  # varsayılan
if 'durumu' in cm:
    st.subheader('Ölüm Değerleri')
    unique_durs = sorted(
        {str(v).strip() for v in df_raw[cm['durumu']].dropna().tolist() if str(v).strip()}
    )
    auto_olum = [v for v in unique_durs if is_olum(v)]

    sess_key = f'olum_values_{cm["durumu"]}'
    if sess_key not in st.session_state:
        # Otomatik öneri sadece geçerli (unique_durs içinde) olanlar
        st.session_state[sess_key] = [v for v in auto_olum if v in unique_durs]

    selected_olum = st.multiselect(
        f'"{cm["durumu"]}" sütununda hangi değerler kuzu ölümünü ifade ediyor?',
        options=unique_durs,
        key=sess_key,
        help='Yıldan yıla "ölüm/öldü/kuzu ölümü" gibi değişen ifadeleri burada elle seç. Otomatik öneri önceden işaretli.',
    )
    olum_set = {str(v).lower().strip().rstrip(':').strip() for v in selected_olum}

    def olum_checker(v, _set=olum_set):
        if pd.isna(v):
            return False
        s = str(v).lower().strip().rstrip(':').strip()
        return s in _set

    if not selected_olum:
        st.warning('Hiç ölüm değeri seçilmedi — yaşama gücü hesabında "ölüm" yokmuş gibi davranılacak.')
    else:
        st.caption(f"✓ Seçili: {len(selected_olum)} değer — {selected_olum}")

# Akıllı uyarı: Grup No eşleşti ama Erken Kızgınlık kaynağı 'grup_no' değil
if 'grup_no' in cm and st.session_state.params.get('kizginlik_kaynak') != 'grup_no':
    c_info, c_btn = st.columns([4, 1])
    c_info.info(
        f"💡 **'Grup No'** alanı `{cm['grup_no']}` sütunuyla eşleşti. "
        f"Şu an Erken Kızgınlık kaynağı **{kaynak_labels.get(st.session_state.params.get('kizginlik_kaynak'), '?')}** — "
        f"hazır grup numaralarını kullanmak için soldaki kaynağı değiştir veya yandaki butona bas."
    )
    if c_btn.button('Grup No\'yu kullan', key='use_grup_no'):
        st.session_state.params['kizginlik_kaynak'] = 'grup_no'
        st.rerun()

# Uyarılar
warnings = validate_data(df_raw, cm)
if warnings:
    st.subheader('Uyarılar')
    for w in warnings:
        if w['severity'] == 'high':
            st.error(w['msg'])
        elif w['severity'] == 'medium':
            st.warning(w['msg'])
        else:
            st.info(w['msg'])
else:
    st.success('✓ Sütun ve format kontrolü temiz.')

# Veri kalite raporu (Excel'deki "2-Veri Kalite Kontrol" sayfasının karşılığı)
st.subheader('Veri Kalite Raporu')
kalite = quality_report(df_raw, cm, olum_checker=olum_checker)
if kalite:
    c1, c2, c3, c4 = st.columns(4)
    c1.metric('Toplam Kayıt',     kalite.get('toplam_kayit', '—'))
    c2.metric('Benzersiz Koyun',  kalite.get('benzersiz_koyun', '—'))
    c3.metric('Benzersiz Kuzu',   kalite.get('benzersiz_kuzu', '—'))
    c4.metric('Ölü Doğum',        kalite.get('olu_dogum', '—'))

    c1, c2, c3, c4 = st.columns(4)
    c1.metric('Tekiz (T)',  kalite.get('dt_T', 0))
    c2.metric('İkiz (İ)',   kalite.get('dt_İ', 0))
    c3.metric('Üçüz (Ü)',   kalite.get('dt_Ü', 0))
    c4.metric('Dördüz (4)', kalite.get('dt_4', 0))

    c1, c2, c3 = st.columns(3)
    c1.metric('Toplam Ölüm',        kalite.get('toplam_olum', 0))
    c2.metric('Ölüm <90 gün',       kalite.get('olum_90_oncesi', 0))
    c3.metric('Ölüm ≥90 gün',       kalite.get('olum_90_sonrasi', 0))

    with st.expander('Eksik Veri Sayıları', expanded=False):
        eksik_data = pd.DataFrame([
            ('Koç No', kalite.get('eksik_koc', '—')),
            ('Koyun No (kritik)', kalite.get('eksik_koyun', '—')),
            ('Doğum Ağırlığı', kalite.get('eksik_dogag', '—')),
            ('Canlı Ağırlık', kalite.get('eksik_canag', '—')),
            ('Ölçüm Tarihi', kalite.get('eksik_olctar', '—')),
        ], columns=['Alan', 'Eksik Sayısı'])
        st.dataframe(eksik_data, use_container_width=True, hide_index=True)

# Devam onayı
high_warn = any(w['severity'] == 'high' for w in warnings)
if 'koyun_no' not in cm or 'dogum_tipi' not in cm:
    st.error('Zorunlu sütunlar eksik (Koyun No, Doğum Tipi). Hesaplama yapılamaz.')
    st.stop()

if high_warn:
    devam = st.checkbox('Yüksek seviyeli uyarılara rağmen devam et', value=False)
    if not devam:
        st.warning('Onay için yukarıdaki kutucuğu işaretle.')
        st.stop()

# ──────────────────────────────────────────────────────────────────────────
# ADIM 3 — FENOTİPİK İNDEKS HESAPLA
# ──────────────────────────────────────────────────────────────────────────
st.header('3. Fenotipik İndeks Hesapla')

if not st.button('Hesapla', type='primary'):
    st.info('Hesaplamayı başlatmak için butona tıkla.')
    st.stop()

with st.spinner('Hesaplanıyor...'):
    df_clean = clean_data(df_raw, cm, P=P, olum_checker=olum_checker)
    piv = build_pivot(df_clean, cm, olum_checker=olum_checker)
    result = calc_index(piv, P)

st.success(f'✓ {len(result)} koyun için fenotipik indeks hesaplandı.')

# Özet istatistikler
c1, c2, c3, c4 = st.columns(4)
c1.metric('Koyun Sayısı', len(result))
c2.metric('Ort. Fenotipik İndeks', f"{result['Fenotipik İndeks'].mean():.2f}")
c3.metric('Min / Max', f"{result['Fenotipik İndeks'].min():.1f} / {result['Fenotipik İndeks'].max():.1f}")
c4.metric('Std. Sapma', f"{result['Fenotipik İndeks'].std():.2f}")

# Sonuç tablosu — sıralı, filtrelenebilir
st.subheader('Sonuç Tablosu')
st.caption('Tüm ara değerler ve nihai puan dahil. Başlığa tıklayarak sıralayabilir, filtreleyebilirsin.')
result_sorted = result.sort_values('Fenotipik İndeks', ascending=False).reset_index(drop=True)
st.dataframe(result_sorted, use_container_width=True, height=500)

# Dağılım grafiği
with st.expander('Fenotipik İndeks Dağılımı', expanded=False):
    chart_data = result_sorted[['Koyun No', 'Fenotipik İndeks']].set_index('Koyun No')
    st.bar_chart(chart_data, height=300)

# ──────────────────────────────────────────────────────────────────────────
# ADIM 4 — İNDİR
# ──────────────────────────────────────────────────────────────────────────
st.header('4. İndir')

ts = datetime.now().strftime('%Y%m%d_%H%M')
c1, c2 = st.columns(2)

with c1:
    xlsx_bytes = to_excel_bytes(result_sorted, P, kalite)
    st.download_button(
        'Excel olarak indir (.xlsx)',
        data=xlsx_bytes,
        file_name=f'Fenotipik_Indeks_{ts}.xlsx',
        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        type='primary',
    )

with c2:
    csv_bytes = result_sorted.to_csv(index=False, encoding='utf-8-sig').encode('utf-8-sig')
    st.download_button(
        'CSV olarak indir (.csv)',
        data=csv_bytes,
        file_name=f'Fenotipik_Indeks_{ts}.csv',
        mime='text/csv',
    )
